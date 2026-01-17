#!/usr/bin/env python3
"""
Automated Test Cases Generator using Claude API
Author: Created for QA Team
Description: Takes requirement PDF, generates comprehensive test cases using Claude API
"""

import os
import sys
import json
import PyPDF2
from anthropic import Anthropic
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Configuration
CLAUDE_API_KEY = os.getenv('ANTHROPIC_API_KEY')
MODEL_NAME = "claude-sonnet-4-20250514"


def read_pdf(pdf_path):
    """Extract text from PDF file"""
    print(f"📄 Reading PDF: {pdf_path}")
    text = ""
    
    try:
        with open(pdf_path, 'rb') as file:
            pdf_reader = PyPDF2.PdfReader(file)
            num_pages = len(pdf_reader.pages)
            print(f"   Total pages: {num_pages}")
            
            for page_num in range(num_pages):
                page = pdf_reader.pages[page_num]
                text += page.extract_text()
                print(f"   Extracted page {page_num + 1}/{num_pages}")
        
        print(f"✅ Successfully extracted {len(text)} characters")
        return text
    
    except Exception as e:
        print(f"❌ Error reading PDF: {e}")
        sys.exit(1)


def generate_test_cases_with_claude(requirements_text, project_name="Project"):
    """Generate test cases using Claude API"""
    print(f"\n🤖 Calling Claude API to generate test cases...")
    
    if not CLAUDE_API_KEY:
        print("❌ Error: ANTHROPIC_API_KEY environment variable not set!")
        print("   Set it using: export ANTHROPIC_API_KEY='your-api-key'")
        sys.exit(1)
    
    client = Anthropic(api_key=CLAUDE_API_KEY)
    
    prompt = f"""You are a professional QA Test Case writer. Based on the following requirements document, create comprehensive test cases.

REQUIREMENTS DOCUMENT:
{requirements_text}

Please generate detailed test cases in JSON format. Each test case should cover:
1. Functional testing
2. Integration testing
3. UI/UX testing
4. Performance testing
5. Security testing
6. Cross-platform testing (if applicable)
7. Edge cases
8. Regression testing

Return ONLY a JSON array with this exact structure (no markdown, no code blocks):

[
  {{
    "id": "TC_001",
    "module": "Module/Feature Name",
    "title": "Clear, concise test case title",
    "description": "Detailed description of what is being tested",
    "preconditions": "Pre-conditions required before testing (numbered list)",
    "steps": "Step-by-step test execution instructions (numbered)",
    "expected": "Expected results (numbered, matching steps)",
    "priority": "P1 or P2 or P3",
    "type": "Functional/Integration/UI/Performance/Security/Cross-Platform/Edge Case/Regression",
    "platform": "Android/iOS/Both/Web"
  }}
]

Guidelines:
- Generate 40-60 comprehensive test cases
- Cover ALL requirements mentioned in the document
- Use clear, professional language
- Make test cases actionable and specific
- Include positive, negative, and edge cases
- Priority: P1 (Critical), P2 (High), P3 (Medium)
- Test steps should be detailed and numbered
- Expected results should match test steps

CRITICAL: Return ONLY the JSON array. No markdown formatting, no ```json blocks, just pure JSON."""

    try:
        response = client.messages.create(
            model=MODEL_NAME,
            max_tokens=16000,
            temperature=0.3,
            messages=[
                {"role": "user", "content": prompt}
            ]
        )
        
        response_text = response.content[0].text
        print("✅ Received response from Claude")
        
        # Clean response if it has markdown code blocks
        if response_text.strip().startswith('```'):
            response_text = response_text.strip()
            if response_text.startswith('```json'):
                response_text = response_text[7:]
            elif response_text.startswith('```'):
                response_text = response_text[3:]
            if response_text.endswith('```'):
                response_text = response_text[:-3]
            response_text = response_text.strip()
        
        # Parse JSON
        test_cases = json.loads(response_text)
        print(f"✅ Successfully parsed {len(test_cases)} test cases")
        return test_cases
    
    except json.JSONDecodeError as e:
        print(f"❌ Error parsing JSON: {e}")
        print(f"Response text: {response_text[:500]}...")
        sys.exit(1)
    
    except Exception as e:
        print(f"❌ Error calling Claude API: {e}")
        sys.exit(1)


def create_excel_file(test_cases, output_path, project_name):
    """Create Excel file with test cases"""
    print(f"\n📝 Creating Excel file...")
    
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Test Cases"
    
    # Styling
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = [
        'Test Case ID',
        'Module',
        'Test Case Title',
        'Description',
        'Pre-conditions',
        'Test Steps',
        'Expected Results',
        'Priority',
        'Test Type',
        'Platform'
    ]
    
    # Column widths
    column_widths = [15, 25, 35, 40, 30, 50, 50, 10, 15, 12]
    for i, width in enumerate(column_widths, 1):
        sheet.column_dimensions[get_column_letter(i)].width = width
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Freeze first row
    sheet.freeze_panes = 'A2'
    
    # Write test cases
    for idx, tc in enumerate(test_cases, start=2):
        row = idx
        sheet[f'A{row}'] = tc.get('id', f'TC_{idx-1:03d}')
        sheet[f'B{row}'] = tc.get('module', '')
        sheet[f'C{row}'] = tc.get('title', '')
        sheet[f'D{row}'] = tc.get('description', '')
        sheet[f'E{row}'] = tc.get('preconditions', '')
        sheet[f'F{row}'] = tc.get('steps', '')
        sheet[f'G{row}'] = tc.get('expected', '')
        sheet[f'H{row}'] = tc.get('priority', 'P2')
        sheet[f'I{row}'] = tc.get('type', 'Functional')
        sheet[f'J{row}'] = tc.get('platform', 'Both')
        
        # Apply borders and wrapping
        for col in range(1, 11):
            cell = sheet.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    # Save
    wb.save(output_path)
    print(f"✅ Excel file created: {output_path}")
    print(f"   Total test cases: {len(test_cases)}")


def generate_summary_stats(test_cases):
    """Generate statistics about test cases"""
    print("\n📊 Test Cases Summary:")
    print("="*60)
    
    # Count by priority
    priorities = {}
    for tc in test_cases:
        priority = tc.get('priority', 'P2')
        priorities[priority] = priorities.get(priority, 0) + 1
    
    print("Priority Distribution:")
    for priority in sorted(priorities.keys()):
        print(f"   {priority}: {priorities[priority]} test cases")
    
    # Count by type
    types = {}
    for tc in test_cases:
        test_type = tc.get('type', 'Functional')
        types[test_type] = types.get(test_type, 0) + 1
    
    print("\nTest Type Distribution:")
    for test_type in sorted(types.keys()):
        print(f"   {test_type}: {types[test_type]} test cases")
    
    # Count by module
    modules = {}
    for tc in test_cases:
        module = tc.get('module', 'General')
        modules[module] = modules.get(module, 0) + 1
    
    print("\nModule Coverage:")
    for module in sorted(modules.keys()):
        print(f"   {module}: {modules[module]} test cases")
    
    print("="*60)


def main():
    """Main function"""
    print("="*60)
    print("🚀 Automated Test Cases Generator using Claude API")
    print("="*60)
    
    # Check arguments
    if len(sys.argv) < 2:
        print("\n❌ Usage: python3 generate_test_cases.py <requirements_pdf_path> [project_name]")
        print("\nExample:")
        print("  python3 generate_test_cases.py requirements.pdf \"My Project\"")
        sys.exit(1)
    
    pdf_path = sys.argv[1]
    project_name = sys.argv[2] if len(sys.argv) > 2 else "Project"
    
    # Check if PDF exists
    if not os.path.exists(pdf_path):
        print(f"❌ Error: PDF file not found: {pdf_path}")
        sys.exit(1)
    
    # Step 1: Read PDF
    requirements_text = read_pdf(pdf_path)
    
    # Step 2: Generate test cases using Claude
    test_cases = generate_test_cases_with_claude(requirements_text, project_name)
    
    # Step 3: Create Excel file
    output_xlsx = f"{project_name.replace(' ', '_')}_Test_Cases.xlsx"
    create_excel_file(test_cases, output_xlsx, project_name)
    
    # Step 4: Generate summary
    generate_summary_stats(test_cases)
    
    # Save JSON for reference
    json_output = f"{project_name.replace(' ', '_')}_Test_Cases.json"
    with open(json_output, 'w', encoding='utf-8') as f:
        json.dump(test_cases, f, indent=2, ensure_ascii=False)
    print(f"✅ JSON saved: {json_output}")
    
    print("\n" + "="*60)
    print("✅ Test Cases Generation Complete!")
    print(f"📄 Excel File: {output_xlsx}")
    print(f"📋 JSON File: {json_output}")
    print("="*60)


if __name__ == "__main__":
    main()
